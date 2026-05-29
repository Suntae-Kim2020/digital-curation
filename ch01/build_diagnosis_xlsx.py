# -*- coding: utf-8 -*-
"""
Chapter 1.4.1 진단 워크시트 Excel 빌더
"""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "데이터 진단"

# 스타일
KOR_FONT = "맑은 고딕"
title_font = Font(name=KOR_FONT, size=14, bold=True, color="FFFFFF")
section_font = Font(name=KOR_FONT, size=11, bold=True, color="1F3A68")
label_font = Font(name=KOR_FONT, size=10)
input_font = Font(name=KOR_FONT, size=10, color="0563C1")
score_font = Font(name=KOR_FONT, size=10, bold=True)

title_fill = PatternFill("solid", fgColor="1F3A68")
section_fill = PatternFill("solid", fgColor="EAF2FB")
input_fill = PatternFill("solid", fgColor="FFFEF0")
total_fill = PatternFill("solid", fgColor="FFF4CE")

thin = Side(style="thin", color="BBBBBB")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

left = Alignment(horizontal="left", vertical="center", wrap_text=True)
center = Alignment(horizontal="center", vertical="center")

# 열 너비
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 28
ws.column_dimensions["C"].width = 40
ws.column_dimensions["D"].width = 10

row = 1

# 제목
ws.cell(row=row, column=1, value="우리 기관 데이터 현황 진단표")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws.cell(row=row, column=1)
c.font = title_font
c.fill = title_fill
c.alignment = center
ws.row_dimensions[row].height = 30
row += 1

# 부제
ws.cell(row=row, column=1, value="『AI 레디 데이터와 디지털 큐레이션』 Chapter 1.4 실습 워크시트")
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
c = ws.cell(row=row, column=1)
c.font = Font(name=KOR_FONT, size=9, italic=True, color="666666")
c.alignment = center
row += 2

# ---------- A. 데이터셋 개요 ----------
ws.cell(row=row, column=1, value="A.").font = section_font
ws.cell(row=row, column=2, value="데이터셋 개요").font = section_font
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = section_fill
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

a_items = [
    "1. 데이터셋 이름",
    "2. 보유 기관·부서",
    "3. 자료 유형 (택1: 도서/논문/보고서/공공기록/멀티미디어/기타)",
    "4. 대략적 규모 (건수)",
    "5. 주된 출처",
]
for item in a_items:
    ws.cell(row=row, column=2, value=item).font = label_font
    ws.cell(row=row, column=2).alignment = left
    ws.cell(row=row, column=2).border = border
    c = ws.cell(row=row, column=3, value="")
    c.fill = input_fill
    c.border = border
    c.alignment = left
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    row += 1

row += 1

# ---------- B. 5가지 조건 자가진단 ----------
ws.cell(row=row, column=1, value="B.").font = section_font
ws.cell(row=row, column=2, value="5가지 조건 자가진단 (각 0~3점)").font = section_font
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = section_fill
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

b_items = [
    ("① 구조화 (Structured)", "스키마 문서 있음=3, 일부=2, 없음=0"),
    ("② 정제 (Clean)", "결측·중복·노이즈 점검 정도"),
    ("③ 식별 (Identified)", "모두 ID 있음=3, 일부=2, 없음=0"),
    ("④ 라이선스 (Licensed)", "표준 표기(CC·공공누리 등) 명시 정도"),
    ("⑤ 의미 부여 (Semantic)", "요약·키워드·임베딩 보유 여부"),
]
score_start_row = row
for label, desc in b_items:
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = left
    ws.cell(row=row, column=2).border = border

    ws.cell(row=row, column=3, value=desc).font = Font(name=KOR_FONT, size=9, color="666666")
    ws.cell(row=row, column=3).alignment = left
    ws.cell(row=row, column=3).border = border

    c = ws.cell(row=row, column=4, value="")
    c.fill = input_fill
    c.font = score_font
    c.alignment = center
    c.border = border
    row += 1
score_end_row = row - 1

# 합계
ws.cell(row=row, column=2, value="합계 / 15").font = Font(name=KOR_FONT, size=11, bold=True, color="1F3A68")
ws.cell(row=row, column=2).alignment = Alignment(horizontal="right", vertical="center")
ws.cell(row=row, column=2).fill = total_fill
ws.cell(row=row, column=2).border = border
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
ws.cell(row=row, column=3).fill = total_fill
ws.cell(row=row, column=3).border = border

c = ws.cell(
    row=row,
    column=4,
    value=f"=SUM(D{score_start_row}:D{score_end_row})",
)
c.font = Font(name=KOR_FONT, size=12, bold=True, color="7A5300")
c.fill = total_fill
c.alignment = center
c.border = border
row += 2

# ---------- C. 가장 약한 조건과 그 이유 ----------
ws.cell(row=row, column=1, value="C.").font = section_font
ws.cell(row=row, column=2, value="가장 약한 조건과 그 이유").font = section_font
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = section_fill
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

for label in ["약한 조건", "이유"]:
    ws.cell(row=row, column=2, value=label).font = label_font
    ws.cell(row=row, column=2).alignment = left
    ws.cell(row=row, column=2).border = border
    c = ws.cell(row=row, column=3, value="")
    c.fill = input_fill
    c.border = border
    c.alignment = left
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    if label == "이유":
        ws.row_dimensions[row].height = 50
    row += 1

row += 1

# ---------- D. AI로 가능해질 활용 시나리오 ----------
ws.cell(row=row, column=1, value="D.").font = section_font
ws.cell(row=row, column=2, value="AI로 가능해질 활용 시나리오 (1~2가지)").font = section_font
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = section_fill
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

c = ws.cell(row=row, column=2, value="")
c.fill = input_fill
c.border = border
c.alignment = left
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
ws.row_dimensions[row].height = 60
row += 2

# ---------- E. 다음 3개월 안에 시도해 볼 한 가지 ----------
ws.cell(row=row, column=1, value="E.").font = section_font
ws.cell(row=row, column=2, value="다음 3개월 안에 시도해 볼 한 가지").font = section_font
for col in range(1, 5):
    ws.cell(row=row, column=col).fill = section_fill
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
row += 1

c = ws.cell(row=row, column=2, value="")
c.fill = input_fill
c.border = border
c.alignment = left
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
ws.row_dimensions[row].height = 50
row += 2

# 안내
note = ws.cell(
    row=row,
    column=1,
    value=(
        "💡 합계 7점 이하 → Ch.2~4 집중 학습 권장 | "
        "10점 이상 → Ch.5~7 집중 학습 권장 | "
        "본 워크시트는 교재 Chapter 1.4.1 실습용입니다."
    ),
)
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
note.font = Font(name=KOR_FONT, size=9, italic=True, color="666666")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[row].height = 30

wb.save("ch01_diagnosis.xlsx")
print("Saved: ch01_diagnosis.xlsx")
